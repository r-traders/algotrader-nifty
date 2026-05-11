"""
Backtest Report Generator
Produces: AlgoTrader_Backtest_Report.pdf  +  AlgoTrader_Backtest.xlsx
"""

import os
import sys
import numpy as np
from datetime import datetime
from typing import Dict, List

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ─────────────────────────────────────────────
# PDF REPORT
# ─────────────────────────────────────────────

def generate_pdf_report(results, config, output_dir: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    BLUE   = colors.HexColor("#58a6ff")
    GREEN  = colors.HexColor("#3fb950")
    RED    = colors.HexColor("#f85149")
    YELLOW = colors.HexColor("#d29922")
    GREY   = colors.HexColor("#8b949e")
    WHITE  = colors.HexColor("#e6edf3")
    DARK   = colors.HexColor("#0d1117")
    PANEL  = colors.HexColor("#161b22")
    BORDER = colors.HexColor("#30363d")
    ORANGE = colors.HexColor("#f0883e")

    def S(name, **kw): return ParagraphStyle(name, **kw)

    TITLE  = S("T", fontSize=26, textColor=BLUE, fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)
    SUB    = S("S", fontSize=11, textColor=GREY, fontName="Helvetica",      alignment=TA_CENTER, spaceAfter=3)
    H1     = S("H1", fontSize=15, textColor=BLUE, fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=6)
    H2     = S("H2", fontSize=11, textColor=GREEN, fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=4)
    BODY   = S("B", fontSize=9, textColor=WHITE, fontName="Helvetica", spaceAfter=4, leading=14)
    NOTE   = S("N", fontSize=8.5, textColor=YELLOW, fontName="Helvetica-Oblique", spaceAfter=4, leftIndent=8)
    CAP    = S("C", fontSize=8, textColor=GREY, fontName="Helvetica", alignment=TA_CENTER)

    def hr(): return HRFlowable(width="100%", thickness=0.5, color=BORDER, spaceAfter=6, spaceBefore=4)
    def sp(h=6): return Spacer(1, h)

    def styled_table(rows, widths, header_color=BLUE):
        t = Table(rows, colWidths=widths)
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0),  PANEL),
            ("TEXTCOLOR",    (0,0), (-1,0),  header_color),
            ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8.5),
            ("BACKGROUND",   (0,1), (-1,-1), DARK),
            ("TEXTCOLOR",    (0,1), (-1,-1), WHITE),
            ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
            ("GRID",         (0,0), (-1,-1), 0.4, BORDER),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [PANEL, DARK]),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ]))
        return t

    out_file = os.path.join(output_dir, "AlgoTrader_Backtest_Report.pdf")
    doc = SimpleDocTemplate(out_file, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm, topMargin=20*mm, bottomMargin=20*mm)

    story = []
    total_pnl   = sum(r.total_pnl for r in results.values())
    total_trades = sum(r.total_trades for r in results.values())
    all_wins     = sum(len(r.winners) for r in results.values())
    overall_wr   = all_wins / total_trades * 100 if total_trades > 0 else 0
    final_cap    = config.initial_capital + total_pnl
    ret_pct      = total_pnl / config.initial_capital * 100

    # COVER
    story += [sp(30), Paragraph("📊 AlgoTrader", TITLE),
              Paragraph("Backtest Report — Jan 2025 to Apr 2, 2026", SUB),
              Paragraph("NSE India | Dhan API | Options · Futures · Equity", SUB),
              sp(16), hr()]

    cover = [
        ["Period",          f"{config.start_date}  →  {config.end_date}"],
        ["Symbols Tested",  ", ".join(results.keys())],
        ["Starting Capital",f"₹{config.initial_capital:,.0f}"],
        ["Final Capital",   f"₹{final_cap:,.0f}"],
        ["Total Return",    f"{ret_pct:+.1f}%"],
        ["Total Trades",    str(total_trades)],
        ["Overall Win Rate",f"{overall_wr:.1f}%"],
        ["Candle Interval", f"{config.interval}-minute"],
        ["Slippage",        f"{config.slippage_pct}% per trade"],
        ["Brokerage",       f"₹{config.brokerage_per_trade} per side"],
        ["Min Confidence",  f"{config.min_confidence}%"],
        ["Min R:R",         f"{config.min_rr}x"],
        ["Report Date",     datetime.now().strftime("%d %b %Y, %H:%M")],
    ]
    story.append(styled_table(cover, [60*mm, 108*mm]))
    story += [sp(16), hr(),
              Paragraph("Prepared for: Rahul Kapoor  |  AlgoTrader v1.0", CAP),
              PageBreak()]

    # OVERALL PERFORMANCE
    story += [Paragraph("1. Overall Performance Summary", H1), hr()]

    overall_rows = [["Symbol", "Trades", "Win%", "Net P&L", "Avg Win", "Avg Loss", "Profit Factor", "Max DD%", "Sharpe"]]
    for sym, r in results.items():
        pnl_color = "+" if r.total_pnl >= 0 else ""
        overall_rows.append([
            sym,
            str(r.total_trades),
            f"{r.win_rate:.1f}%",
            f"₹{r.total_pnl:+,.0f}",
            f"₹{r.avg_win:,.0f}",
            f"₹{r.avg_loss:,.0f}",
            str(r.profit_factor),
            f"{r.max_drawdown:.1f}%",
            str(r.sharpe_ratio),
        ])
    # Total row
    overall_rows.append([
        "TOTAL",
        str(total_trades),
        f"{overall_wr:.1f}%",
        f"₹{total_pnl:+,.0f}",
        "", "", "", "", ""
    ])

    ot = Table(overall_rows, colWidths=[24*mm, 14*mm, 14*mm, 22*mm, 18*mm, 18*mm, 22*mm, 16*mm, 16*mm])
    ot.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  PANEL),
        ("TEXTCOLOR",    (0,0), (-1,0),  BLUE),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("BACKGROUND",   (0,1), (-1,-2), DARK),
        ("BACKGROUND",   (0,-1),(-1,-1), PANEL),
        ("TEXTCOLOR",    (0,1), (0,-2),  GREEN),
        ("TEXTCOLOR",    (0,-1),(-1,-1), YELLOW),
        ("FONTNAME",     (0,-1),(-1,-1), "Helvetica-Bold"),
        ("TEXTCOLOR",    (1,1), (-1,-2), WHITE),
        ("FONTNAME",     (0,1), (-1,-2), "Helvetica"),
        ("GRID",         (0,0), (-1,-1), 0.4, BORDER),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [PANEL, DARK]),
        ("LEFTPADDING",  (0,0), (-1,-1), 6),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("ALIGN",        (1,0), (-1,-1), "CENTER"),
    ]))
    story.append(ot)
    story.append(sp(8))

    # Key metrics boxes (labels row + values row)
    metrics_labels = ["Starting Capital", "Final Capital", "Net Profit/Loss", "Total Return"]
    metrics_values = [
        f"₹{config.initial_capital:,.0f}",
        f"₹{final_cap:,.0f}",
        f"₹{total_pnl:+,.0f}",
        f"{ret_pct:+.1f}%",
    ]
    mt = Table([metrics_labels, metrics_values], colWidths=[42*mm]*4)
    mt.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), PANEL),
        ("TEXTCOLOR",    (0,0), (-1,0),  GREY),
        ("TEXTCOLOR",    (0,1), (-1,1),  WHITE),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica"),
        ("FONTNAME",     (0,1), (-1,1),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("GRID",         (0,0), (-1,-1), 0.5, BORDER),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
    ]))
    story.append(mt)
    story.append(PageBreak())

    # PER SYMBOL BREAKDOWN
    story += [Paragraph("2. Per-Symbol Breakdown", H1), hr()]

    for sym, r in results.items():
        story.append(Paragraph(f"2.{list(results.keys()).index(sym)+1}  {sym}", H2))
        sym_rows = [
            ["Total Trades",          str(r.total_trades),      "Profit Factor",       str(r.profit_factor)],
            ["Winners",               str(len(r.winners)),      "Max Drawdown",        f"{r.max_drawdown:.1f}%"],
            ["Losers",                str(len(r.losers)),        "Sharpe Ratio",        str(r.sharpe_ratio)],
            ["Win Rate",              f"{r.win_rate:.1f}%",     "Max Consec Losses",   str(r.max_consecutive_losses)],
            ["Average Win",           f"₹{r.avg_win:,.0f}",    "Net P&L",             f"₹{r.total_pnl:+,.0f}"],
            ["Average Loss",          f"₹{r.avg_loss:,.0f}",   "Return on Capital",   f"{r.total_pnl/config.initial_capital*100:+.1f}%"],
        ]
        st = Table(sym_rows, colWidths=[44*mm, 28*mm, 44*mm, 52*mm])
        st.setStyle(TableStyle([
            ("FONTSIZE",     (0,0), (-1,-1), 8.5),
            ("BACKGROUND",   (0,0), (-1,-1), DARK),
            ("TEXTCOLOR",    (0,0), (0,-1),  GREY),
            ("TEXTCOLOR",    (2,0), (2,-1),  GREY),
            ("TEXTCOLOR",    (1,0), (1,-1),  WHITE),
            ("TEXTCOLOR",    (3,0), (3,-1),  WHITE),
            ("FONTNAME",     (0,0), (-1,-1), "Helvetica"),
            ("GRID",         (0,0), (-1,-1), 0.4, BORDER),
            ("ROWBACKGROUNDS",(0,0),(-1,-1), [PANEL, DARK]),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ]))
        story.append(st)
        story.append(sp(10))

        # Top 5 trades
        if r.trades:
            story.append(Paragraph(f"Top 5 Winning Trades — {sym}", S("th", fontSize=9, textColor=GREEN,
                fontName="Helvetica-Bold", spaceAfter=4)))
            top_wins = sorted(r.trades, key=lambda t: t.pnl_net, reverse=True)[:5]
            tw_rows = [["Date", "Dir", "Entry", "Exit", "Qty", "P&L Net", "Reason"]]
            for t in top_wins:
                tw_rows.append([t.entry_date, t.direction,
                    f"₹{t.entry_price:.2f}", f"₹{t.exit_price:.2f}",
                    str(t.quantity), f"₹{t.pnl_net:+,.0f}", t.exit_reason])
            story.append(styled_table(tw_rows, [24*mm, 12*mm, 22*mm, 22*mm, 12*mm, 22*mm, 30*mm], GREEN))
            story.append(sp(6))

    story.append(PageBreak())

    # RISK NOTES
    story += [Paragraph("3. Risk & Disclaimer", H1), hr(),
              Paragraph("This backtest was conducted using historical price data from NSE via Dhan API. "
                        "Results include realistic slippage (0.05%) and brokerage (₹40/side). "
                        "Option chain signals were not available in the backtest (historical OC data not in scope) — "
                        "only Indicator and SMC signals were used. Live results may differ significantly.", BODY),
              sp(6),
              Paragraph("⚠ Past performance does not guarantee future results. "
                        "Always run in paper trading mode for at least 2 weeks before going live. "
                        "Never risk more than you can afford to lose.", NOTE),
              sp(16), hr(),
              Paragraph(f"AlgoTrader Backtest Report  |  Rahul Kapoor  |  {datetime.now().strftime('%d %b %Y')}", CAP)]

    doc.build(story)
    return out_file


# ─────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────

def generate_excel_report(results, config, all_trades: List[dict], output_dir: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, Reference, BarChart
    except ImportError:
        os.system(f"{sys.executable} -m pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, Reference, BarChart

    wb = openpyxl.Workbook()

    DARK_FILL  = PatternFill("solid", fgColor="0D1117")
    PANEL_FILL = PatternFill("solid", fgColor="161B22")
    BLUE_FILL  = PatternFill("solid", fgColor="1C2D4A")
    GREEN_FILL = PatternFill("solid", fgColor="1A4731")
    RED_FILL   = PatternFill("solid", fgColor="3D1F1F")
    HEADER_FONT = Font(name="Calibri", bold=True, color="58A6FF", size=10)
    GREEN_FONT  = Font(name="Calibri", color="3FB950", size=10)
    RED_FONT    = Font(name="Calibri", color="F85149", size=10)
    WHITE_FONT  = Font(name="Calibri", color="E6EDF3", size=10)
    YELLOW_FONT = Font(name="Calibri", color="D29922", size=10, bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")

    def style_header_row(ws, row, fill=PANEL_FILL):
        for cell in ws[row]:
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    def style_data_row(ws, row_num, green=False, red=False):
        fill = GREEN_FILL if green else RED_FILL if red else DARK_FILL
        for cell in ws[row_num]:
            cell.fill = fill
            cell.font = GREEN_FONT if green else RED_FONT if red else WHITE_FONT
            cell.alignment = Alignment(vertical="center")

    # ── SHEET 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "58A6FF"
    ws.sheet_view.showGridLines = False

    ws.append(["AlgoTrader — Backtest Report", "", "Jan 2025 – Apr 2, 2026"])
    ws.append([""])
    ws.append(["Symbol", "Trades", "Win%", "Net P&L (₹)", "Avg Win (₹)", "Avg Loss (₹)",
               "Profit Factor", "Max Drawdown%", "Sharpe Ratio", "Max Consec Loss"])
    style_header_row(ws, 3)

    for sym, r in results.items():
        ws.append([sym, r.total_trades, round(r.win_rate,1), round(r.total_pnl,2),
                   round(r.avg_win,2), round(r.avg_loss,2), r.profit_factor,
                   r.max_drawdown, r.sharpe_ratio, r.max_consecutive_losses])
        style_data_row(ws, ws.max_row, green=r.total_pnl > 0, red=r.total_pnl <= 0)

    total_pnl = sum(r.total_pnl for r in results.values())
    ws.append(["TOTAL", sum(r.total_trades for r in results.values()),
               "", round(total_pnl, 2), "", "", "", "", "", ""])
    for cell in ws[ws.max_row]:
        cell.fill = PANEL_FILL
        cell.font = YELLOW_FONT
        cell.alignment = CENTER

    # Column widths
    for i, w in enumerate([16,10,10,16,14,14,14,16,14,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── SHEET 2: All Trades ──
    ws2 = wb.create_sheet("All Trades")
    ws2.sheet_view.showGridLines = False

    headers = ["#", "Symbol", "Direction", "Type", "Entry Date", "Entry Time",
               "Entry Price", "Exit Date", "Exit Time", "Exit Price",
               "Qty", "SL", "Target", "P&L Gross", "P&L Net", "P&L%",
               "Exit Reason", "Confidence", "R:R", "Bars Held"]
    ws2.append(headers)
    style_header_row(ws2, 1)

    for t in all_trades:
        row = [
            t.get("trade_id"), t.get("symbol"), t.get("direction"), t.get("instrument_type"),
            t.get("entry_date"), t.get("entry_time"), t.get("entry_price"),
            t.get("exit_date"), t.get("exit_time"), t.get("exit_price"),
            t.get("quantity"), t.get("stop_loss"), t.get("target_1"),
            t.get("pnl_gross"), t.get("pnl_net"), t.get("pnl_pct"),
            t.get("exit_reason"), t.get("confidence"), t.get("rr_actual"), t.get("bars_held"),
        ]
        ws2.append(row)
        pnl = float(t.get("pnl_net", 0))
        style_data_row(ws2, ws2.max_row, green=pnl > 0, red=pnl <= 0)

    for i, w in enumerate([6,12,10,8,12,10,13,12,10,13,6,13,13,13,13,8,14,10,6,10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── SHEET 3: Equity Curves ──
    ws3 = wb.create_sheet("Equity Curves")
    ws3.sheet_view.showGridLines = False

    symbols = list(results.keys())
    headers3 = ["Bar"] + symbols
    ws3.append(headers3)
    style_header_row(ws3, 1)

    max_len = max(len(r.equity_curve) for r in results.values())
    for i in range(max_len):
        row3 = [i + 1]
        for sym in symbols:
            eq = results[sym].equity_curve
            row3.append(eq[i] if i < len(eq) else "")
        ws3.append(row3)
        for cell in ws3[ws3.max_row]:
            cell.fill = DARK_FILL
            cell.font = WHITE_FONT

    # Add equity chart
    chart = LineChart()
    chart.title = "Equity Curves — All Symbols"
    chart.style = 10
    chart.width = 22
    chart.height = 14
    data_ref = Reference(ws3, min_col=2, max_col=len(symbols)+1, min_row=1, max_row=min(max_len+1, 500))
    chart.add_data(data_ref, titles_from_data=True)
    ws3.add_chart(chart, "A" + str(max_len + 5))

    # ── SHEET 4: Monthly P&L ──
    ws4 = wb.create_sheet("Monthly P&L")
    ws4.sheet_view.showGridLines = False
    ws4.append(["Month"] + symbols + ["Total"])
    style_header_row(ws4, 1)

    # Aggregate by month
    monthly: Dict[str, Dict[str, float]] = {}
    for sym, r in results.items():
        for d, pnl in r.daily_pnl.items():
            month = d[:7]  # YYYY-MM
            if month not in monthly:
                monthly[month] = {s: 0.0 for s in symbols}
            monthly[month][sym] = monthly[month].get(sym, 0.0) + pnl

    for month in sorted(monthly.keys()):
        row4 = [month] + [round(monthly[month].get(s, 0), 2) for s in symbols]
        total_month = sum(monthly[month].get(s, 0) for s in symbols)
        row4.append(round(total_month, 2))
        ws4.append(row4)
        style_data_row(ws4, ws4.max_row, green=total_month > 0, red=total_month <= 0)

    for i in range(1, len(symbols) + 3):
        ws4.column_dimensions[get_column_letter(i)].width = 14

    out_file = os.path.join(output_dir, "AlgoTrader_Backtest.xlsx")
    wb.save(out_file)
    return out_file
